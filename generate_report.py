from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from supabase import create_client, Client
import os
from io import BytesIO
from PIL import Image as PILImage
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

url = os.getenv("SUPABASE_URL")
key = os.getenv("SUPABASE_KEY")
supabase: Client = create_client(url, key)

def generate_attendance_excel(session_data, students):
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"

    # Add School Logo (Ensure logo exists in assets folder)
    logo_path = "assets/logo.png"
    if os.path.exists(logo_path):
        # Resize the logo proportionally to a height of 60px
        logo = PILImage.open(logo_path)
        aspect_ratio = logo.width / logo.height
        new_height = 60
        new_width = int(aspect_ratio * new_height)
        logo = logo.resize((new_width, new_height))

        # Save resized logo temporarily in memory
        logo_bytes = BytesIO()
        logo.save(logo_bytes, format="PNG")
        logo_bytes.seek(0)

        # Insert resized logo into Excel
        img = Image(logo_bytes)
        img.anchor = 'B1'  # Position the logo (column B, row 1)
        ws.add_image(img)
    else:
        print("Logo not found at assets/logo.png")

    # Add Session Details (Shifted by 3 rows)
    ws['C6'] = f"Attendance Report for {session_data['specialty']} - {session_data['group']}"
    ws['C7'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['C8'] = f"Instructor: {session_data.get('teacher_name', 'Unknown')}"

    # Table Headers (Shifted by 3 rows)
    ws['A10'] = "No"
    ws['B10'] = "Student ID"
    ws['C10'] = "Name"
    ws['D10'] = "Status"

    # Populate Student Data (Start from row 11)
    for idx, student in enumerate(students, start=1):
        ws[f'A{10 + idx}'] = idx
        ws[f'B{10 + idx}'] = student['id']
        ws[f'C{10 + idx}'] = f"{student['first_name']} {student['last_name']}"
        ws[f'D{10 + idx}'] = student['status']

    # Save workbook to memory (not disk)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Upload directly to Supabase
    file_name = f"{session_data['session_id']}.xlsx"
    supabase.storage.from_('sessions').upload(file_name, output.read())

    print(f"Attendance uploaded directly to Supabase as {file_name}")
